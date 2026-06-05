"""
License generator for Siegetower.

Generates a hardware-bound license.dat file for a specific user machine.
Keep this tool private - never distribute it.

Usage:
    python tools/generate_license.py <machine_id> <name> [options]
"""
from __future__ import annotations

import argparse
import base64
import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import padding
except ImportError:
    print('[X] cryptography package not installed. Run: pip install cryptography')
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print('[X] openpyxl package not installed. Run: pip install openpyxl')
    sys.exit(1)

TOOLS_DIR = Path(__file__).parent
PROJECT_ROOT = TOOLS_DIR.parent
PRIVATE_KEY_PATH = PROJECT_ROOT / 'licenses' / 'keys' / 'private_key.pem'
LICENSES_ROOT = PROJECT_ROOT / 'licenses'
LICENSES_DAT_ROOT = LICENSES_ROOT / 'dat'
CONTROL_XLSX_PATH = LICENSES_ROOT / 'control_licenses.xlsx'


def _sanitize_username(value: str) -> str:
    sanitized = re.sub(r'[^A-Za-z0-9._-]+', '_', value.strip())
    sanitized = sanitized.strip('._-')
    return sanitized or 'user'


def _default_output_path(issued_to: str, now: datetime) -> Path:
    month_folder = now.strftime('%Y-%m')
    username = _sanitize_username(issued_to)
    return LICENSES_DAT_ROOT / month_folder / f'{username}_license.dat'


def _update_control_spreadsheet(
    *,
    issued_to: str,
    generated_date: str,
    expiry: str,
    renew_in_days: str,
    machine_id: str,
) -> None:
    LICENSES_ROOT.mkdir(parents=True, exist_ok=True)

    headers = ['User Name', 'Generated Date', 'Expiry Date', 'Renew In (Days)', 'PC Code']
    if CONTROL_XLSX_PATH.exists():
        wb = load_workbook(CONTROL_XLSX_PATH)
        ws = wb.active
        if ws.max_row < 1:
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'licenses'
        ws.append(headers)

    ws.append([issued_to, generated_date, expiry if expiry else 'never', renew_in_days, machine_id])
    wb.save(CONTROL_XLSX_PATH)


def generate_license(
    machine_id: str,
    issued_to: str,
    expiry: str,
    output_path: Path | None,
) -> None:
    if not PRIVATE_KEY_PATH.exists():
        print(f'[X] Private key not found: {PRIVATE_KEY_PATH}')
        print('    Run tools/generate_keys.py first.')
        sys.exit(1)

    private_key_pem = PRIVATE_KEY_PATH.read_bytes()
    private_key = serialization.load_pem_private_key(private_key_pem, password=None)

    payload = json.dumps(
        {'expiry': expiry, 'issued_to': issued_to, 'machine_id': machine_id},
        sort_keys=True,
    ).encode('utf-8')

    signature = private_key.sign(
        payload,
        padding.PSS(
            mgf=padding.MGF1(hashes.SHA256()),
            salt_length=padding.PSS.MAX_LENGTH,
        ),
        hashes.SHA256(),
    )

    license_data = {
        'machine_id': machine_id,
        'issued_to': issued_to,
        'expiry': expiry,
        'signature': base64.b64encode(signature).decode(),
    }

    now = datetime.now()
    final_output = output_path or _default_output_path(issued_to, now)

    final_output.parent.mkdir(parents=True, exist_ok=True)
    final_output.write_text(json.dumps(license_data, indent=2), encoding='utf-8')

    generated_date = now.date().isoformat()
    renew_in_days = 'N/A'
    if expiry:
        try:
            expiry_date = date.fromisoformat(expiry)
            renew_in_days = str((expiry_date - now.date()).days)
        except ValueError:
            renew_in_days = 'INVALID_EXPIRY'

    _update_control_spreadsheet(
        issued_to=issued_to,
        generated_date=generated_date,
        expiry=expiry,
        renew_in_days=renew_in_days,
        machine_id=machine_id,
    )

    expiry_display = expiry if expiry else 'Never'
    print(f'\n[OK] License created: {final_output}')
    print(f'     Machine ID  : {machine_id}')
    print(f'     Issued to   : {issued_to}')
    print(f'     Expires     : {expiry_display}')
    print(f'     Control XLSX: {CONTROL_XLSX_PATH}')
    print('\n[>] Send this license.dat file to the user.')
    print('    They must place it in: %APPDATA%\\Siegetower\\license.dat')


def main() -> None:
    parser = argparse.ArgumentParser(
        description='Generate a Siegetower license for a specific machine.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('machine_id', help='Target machine ID (e.g. A3F2-9C1B-D47E-8801)')
    parser.add_argument('name', help='Licensee name or identifier')
    parser.add_argument(
        '--expiry',
        default=(date.today() + timedelta(days=365)).isoformat(),
        help='Expiry date as YYYY-MM-DD, or "never" for no expiry (default: 1 year from today)',
    )
    parser.add_argument(
        '--output',
        default=None,
        help='Optional output file path override. By default: licenses/dat/YYYY-MM/<username>_license.dat',
    )
    args = parser.parse_args()

    expiry = '' if args.expiry.lower() == 'never' else args.expiry

    generate_license(
        machine_id=args.machine_id.upper().strip(),
        issued_to=args.name,
        expiry=expiry,
        output_path=Path(args.output) if args.output else None,
    )


if __name__ == '__main__':
    main()
